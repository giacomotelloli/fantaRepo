import openpyxl
import csv
import numpy as np
import numbers
from tqdm import tqdm
class Stagione:
    def __init__(self,anno,giornateTrascorse):
        self.anno=anno
        self.giornateTrascorse=giornateTrascorse

class TimeSuccPlayer:
    def __init__(self,ruolo,nomeGioc,sq,gMax):
        self.nome=nomeGioc
        self.squadra=sq
        self.ruolo=ruolo
        self.listaSucc=np.zeros(gMax)

    def addElem(self,number,pos):
        self.listaSucc[pos-1]=number

#generare listone di oggetti timesuccplayer
file = openpyxl.load_workbook('/home/giacomo/Documenti/Acr/Fanta_ACR/Statistiche_Fantacalcio_2021-22.xlsx')
sheet=file.worksheets[0]
giornateCamp=38
stagioni=[Stagione("2019-20",38),Stagione("2020-21",38),Stagione("2021-22",23)]
pesiModificatori=[3,-1,3,-3,3,-2,-0.5,-1,1,1]
giornateTot=0
giornateTot=sum([s.giornateTrascorse for s in stagioni])
print("Giornate totali considerate:"+str(giornateTot))
Gf=3
Gs=-1
Rp=3
Rs=-3
Rf=3
Au=-2
Amm=-0.5
Esp=-1
Ass=+1
Asf=+1
Listone=[]

def aggiungiFantaVoto(nome,voto,ruolo,giornata):
    for i in range(len(Listone)):
        if Listone[i].nome==nome and Listone[i].ruolo==ruolo:
            Listone[i].addElem(voto,giornata)
            break
    return

for row in sheet.iter_rows():
    Listone.append(TimeSuccPlayer(row[1].value, row[2].value,row[3].value,giornateTot))

#ciclo principale per aprire i file uno ad uno(indice i)
#per generare un array per ogni giocatore con la successione dei suoi voti


count=0
for s in stagioni:

    for i in tqdm(range(1,s.giornateTrascorse+1)):
        file = openpyxl.load_workbook('/home/giacomo/Documenti/Acr/Fanta_ACR/voti_time_serie/'+s.anno+'/Voti_Fantacalcio_Stagione_'+s.anno+'_Giornata_'+str(i)+'.xlsx')
        sheet = file.worksheets[0]
# per ogni giocatore nel file calcolo il fantavoto
# poi lo cerco nel listone e quando lo trovo aggiorno la listasucc nella posizione i
        for row in sheet.iter_rows():
            if row[1].value=="P" or row[1].value=="D" or row[1].value=="C" or row[1].value=="A":
                ruolo=row[1].value
                nome=row[2].value
                if isinstance(row[3].value,numbers.Number):
                    voto=row[3].value
                else:
                    voto=row[3].value.split("*")[0]
                    voto=int(voto)
                for j in range(10):#calcolo il fantavoto accedendo ai valori delle colonne
                    voto=voto + (pesiModificatori[j]*row[4+j].value)
                aggiungiFantaVoto(nome,voto,ruolo,i+count)

    count+=38

#alla fine del ciclo principale salvo tutto in un file csv
with open('/home/giacomo/Documenti/Acr/Fanta_ACR/voti_time_serie/saving_names.csv', 'w',newline='') as file:
    writer = csv.writer(file)
    writer.writerow(["Ruolo","Nome", "Squadra"])

    for i in range(len(Listone)):
        writer.writerow([Listone[i].nome,Listone[i].squadra,Listone[i].ruolo])

with open('/home/giacomo/Documenti/Acr/Fanta_ACR/voti_time_serie/saving_marks.csv', 'w',newline='') as file:
    writer = csv.writer(file)

    for i in range(len(Listone)):
        writer.writerow(Listone[i].listaSucc)
